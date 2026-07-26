import streamlit as st

from streamlit_js_eval import streamlit_js_eval

from math import radians, sin, cos, sqrt, atan2

from datetime import datetime

import os

from datetime import date

from openpyxl import Workbook, load_workbook

from openpyxl.styles import PatternFill

from zoneinfo import ZoneInfo

import pandas as pd


# ============================================================
# CONFIGURACIÓN DE LA PÁGINA
# ============================================================

st.set_page_config(
    page_title="Asistencia Ciclo II 2026B",
    page_icon="📋",
    layout="wide"
)


# ============================================================
# CONFIGURACIÓN DEL CAM
# ============================================================

LATITUD_CAM = 4.681785 #4.7556 
LONGITUD_CAM = -74.216075 #-74.0237 

# Radio máximo permitido
RADIO_PERMITIDO = 500  # metros


# ============================================================
# ZONA HORARIA (Bogotá, sin importar en qué país esté el servidor)
# ============================================================

ZONA_BOGOTA = ZoneInfo("America/Bogota")


def ahora_bogota():
    """Hora actual en Bogotá. Se devuelve "naive" (sin tzinfo) porque
    openpyxl no admite guardar fechas con zona horaria en el Excel."""

    return datetime.now(ZONA_BOGOTA).replace(tzinfo=None)


# ============================================================
# CONFIGURACIÓN DEL EXCEL Y DEL HORARIO
# ============================================================

ARCHIVO_EXCEL = "asistencia.xlsx"

# weekday(): lunes=0, martes=1, miércoles=2, jueves=3, viernes=4, sábado=5, domingo=6
DIAS_HORARIO = {1, 4}   # martes y viernes

HORA_INICIO_HORARIO = 7    # 7:00 am
HORA_FIN_HORARIO = 12      # 12:00 m (mediodía)

HORAS_MAX_DIA = 8          # tope de seguridad por día (evita horas absurdas si alguien olvida dar salida)
HORAS_REQUERIDAS = 200     # horas totales exigidas en el ciclo

NOMBRES_DIAS_ES = ["lun", "mar", "mié", "jue", "vie", "sáb", "dom"]
MESES_ES = [
    "ene", "feb", "mar", "abr", "may", "jun",
    "jul", "ago", "sep", "oct", "nov", "dic"
]


def formatear_fecha_columna(fecha):
    """Convierte una fecha en algo como 'vie 17-jul-26'."""

    return (
        f"{NOMBRES_DIAS_ES[fecha.weekday()]} "
        f"{fecha.day:02d}-{MESES_ES[fecha.month - 1]}-{str(fecha.year)[2:]}"
    )


def inicializar_excel():
    """Crea el archivo con sus 3 hojas si no existe, y completa
    cualquier hoja que le falte si el archivo ya existe (por ejemplo,
    archivos creados con una versión anterior de la app)."""

    if not os.path.exists(ARCHIVO_EXCEL):

        wb = Workbook()

        ws_reg = wb.active
        ws_reg.title = "Registros"
        ws_reg.append(["Nombre", "Cultivo", "Ingreso", "Salida", "Horas"])

        ws_dev = wb.create_sheet("_Dispositivos")
        ws_dev.append(["Fecha", "Accion", "DeviceID", "Nombre"])
        ws_dev.sheet_state = "hidden"

        ws_cal = wb.create_sheet("Calendario")
        ws_cal.append(["Estudiante"])

        wb.save(ARCHIVO_EXCEL)

        return

    # El archivo ya existe: revisar que tenga las 3 hojas requeridas
    wb = load_workbook(ARCHIVO_EXCEL)
    modificado = False

    if "Registros" not in wb.sheetnames:
        ws_reg = wb.create_sheet("Registros")
        ws_reg.append(["Nombre", "Cultivo", "Ingreso", "Salida", "Horas"])
        modificado = True

    if "_Dispositivos" not in wb.sheetnames:
        ws_dev = wb.create_sheet("_Dispositivos")
        ws_dev.append(["Fecha", "Accion", "DeviceID", "Nombre"])
        ws_dev.sheet_state = "hidden"
        modificado = True

    if "Calendario" not in wb.sheetnames:
        ws_cal = wb.create_sheet("Calendario")
        ws_cal.append(["Estudiante"])
        modificado = True

    # Migración: si "Registros" existe pero de una versión sin
    # columna "Cultivo", se le agrega (vacía para las filas viejas)
    # sin perder ningún dato existente.
    ws_reg = wb["Registros"]
    encabezados = [c.value for c in ws_reg[1]]

    if "Cultivo" not in encabezados:
        ws_reg.insert_cols(2)
        ws_reg.cell(row=1, column=2, value="Cultivo")
        modificado = True

    if modificado:
        wb.save(ARCHIVO_EXCEL)


def obtener_fila_de_hoy(ws_reg, nombre, fecha_hoy):
    """Busca la fila de Registros de este estudiante para la fecha dada."""

    for i, fila in enumerate(ws_reg.iter_rows(min_row=2, values_only=True), start=2):

        nombre_fila, cultivo_fila, ingreso, salida, horas = fila

        if nombre_fila == nombre and ingreso and ingreso.date() == fecha_hoy:
            return i

    return None


def calcular_horas_dia(ingreso, salida):
    """Devuelve (horas_totales, horas_dentro_horario, horas_fuera_horario) para un día."""

    if not ingreso or not salida:
        return 0.0, 0.0, 0.0

    horas_brutas = (salida - ingreso).total_seconds() / 3600
    horas_totales = max(0.0, min(horas_brutas, HORAS_MAX_DIA))

    fecha_dia = ingreso.date()
    horas_dentro = 0.0

    if fecha_dia.weekday() in DIAS_HORARIO:

        inicio_horario = ingreso.replace(
            hour=HORA_INICIO_HORARIO, minute=0, second=0, microsecond=0
        )

        fin_horario = ingreso.replace(
            hour=HORA_FIN_HORARIO, minute=0, second=0, microsecond=0
        )

        solape_inicio = max(ingreso, inicio_horario)
        solape_fin = min(salida, fin_horario)

        if solape_fin > solape_inicio:
            horas_dentro = (solape_fin - solape_inicio).total_seconds() / 3600

    horas_dentro = min(horas_dentro, horas_totales)
    horas_fuera = horas_totales - horas_dentro

    return horas_totales, horas_dentro, horas_fuera


RELLENO_FUERA_HORARIO = PatternFill(
    start_color="FFD9A6", end_color="FFD9A6", fill_type="solid"
)

RELLENO_DENTRO_HORARIO = PatternFill(
    start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"
)


def recalcular_calendario(wb):
    """Reconstruye por completo la hoja Calendario a partir de Registros.

    Se crea una columna por CADA fecha que tenga al menos un registro,
    sin importar el día de la semana. Las fechas que no son día
    programado (martes/viernes) se resaltan en naranja para diferenciarlas.
    """

    ws_reg = wb["Registros"]

    fechas_con_registro = set()
    datos_por_estudiante = {}

    for nombre, cultivo, ingreso, salida, horas in ws_reg.iter_rows(min_row=2, values_only=True):

        if not nombre or not ingreso:
            continue

        fecha_dia = ingreso.date()

        totales, dentro, fuera = calcular_horas_dia(ingreso, salida)

        fechas_con_registro.add(fecha_dia)

        datos_por_estudiante.setdefault(nombre, {})
        datos_por_estudiante[nombre][fecha_dia] = {
            "totales": totales,
            "dentro": dentro,
            "fuera": fuera
        }

    fechas_ordenadas = sorted(fechas_con_registro)

    if "Calendario" in wb.sheetnames:
        del wb["Calendario"]

    ws_cal = wb.create_sheet("Calendario")

    encabezado = (
        ["Estudiante"]
        + [formatear_fecha_columna(f) for f in fechas_ordenadas]
        + [
            "Horas dentro de horario",
            "Horas fuera de horario",
            "Total horas",
            "% Horario",
            "% Total"
        ]
    )

    ws_cal.append(encabezado)

    # Colorear encabezados de fechas que NO son día programado (martes/viernes)
    for col_idx, fecha in enumerate(fechas_ordenadas, start=2):

        if fecha.weekday() not in DIAS_HORARIO:
            ws_cal.cell(row=1, column=col_idx).fill = RELLENO_FUERA_HORARIO
        else:
            ws_cal.cell(row=1, column=col_idx).fill = RELLENO_DENTRO_HORARIO

    fila_actual = 2

    for nombre in sorted(datos_por_estudiante.keys()):

        fila = [nombre]

        total_dentro = 0.0
        total_fuera = 0.0

        for f in fechas_ordenadas:

            registro_dia = datos_por_estudiante[nombre].get(f)

            if registro_dia:
                fila.append(round(registro_dia["totales"], 2))
                total_dentro += registro_dia["dentro"]
                total_fuera += registro_dia["fuera"]
            else:
                fila.append("")

        total_horas = total_dentro + total_fuera

        pct_horario = (
            (total_dentro / HORAS_REQUERIDAS * 100) if HORAS_REQUERIDAS else 0
        )

        pct_total = (
            (total_horas / HORAS_REQUERIDAS * 100) if HORAS_REQUERIDAS else 0
        )

        fila += [
            round(total_dentro, 2),
            round(total_fuera, 2),
            round(total_horas, 2),
            f"{pct_horario:.1f}%",
            f"{pct_total:.1f}%"
        ]

        ws_cal.append(fila)

        # Colorear también las celdas de datos de las fechas fuera de horario
        for col_idx, fecha in enumerate(fechas_ordenadas, start=2):

            valor_celda = ws_cal.cell(row=fila_actual, column=col_idx).value

            if valor_celda != "":
                if fecha.weekday() not in DIAS_HORARIO:
                    ws_cal.cell(row=fila_actual, column=col_idx).fill = RELLENO_FUERA_HORARIO
                else:
                    ws_cal.cell(row=fila_actual, column=col_idx).fill = RELLENO_DENTRO_HORARIO

        fila_actual += 1


def dispositivo_ya_registro(device_id, fecha, accion, estudiante=None):
    """Revisa si este dispositivo ya registró esta acción en la fecha dada.

    Si se indica 'estudiante', además exige que el registro coincida
    con ese nombre (para evitar que un dispositivo dé entrada a un
    estudiante y salida a otro distinto)."""

    if not device_id:
        return False

    inicializar_excel()

    wb = load_workbook(ARCHIVO_EXCEL)
    ws = wb["_Dispositivos"]

    for fila_fecha, fila_accion, fila_device, fila_nombre in ws.iter_rows(min_row=2, values_only=True):

        if (
            fila_device == device_id
            and fila_fecha == fecha
            and fila_accion == accion
        ):
            if estudiante is None or fila_nombre == estudiante:
                return True

    return False


def guardar_registro(estudiante, cultivo, accion, distancia, device_id, fecha_hora):
    """Guarda la entrada/salida en Registros, registra el dispositivo y recalcula el Calendario."""

    inicializar_excel()

    wb = load_workbook(ARCHIVO_EXCEL)

    ws_reg = wb["Registros"]
    ws_dev = wb["_Dispositivos"]

    fecha_hoy = fecha_hora.date()

    fila_existente = obtener_fila_de_hoy(ws_reg, estudiante, fecha_hoy)

    if accion == "Entrada":

        if fila_existente is None:
            ws_reg.append([estudiante, cultivo, fecha_hora, None, None])
        else:
            ws_reg.cell(row=fila_existente, column=2, value=cultivo)
            ws_reg.cell(row=fila_existente, column=3, value=fecha_hora)

    else:  # Salida

        if fila_existente is not None:

            ws_reg.cell(row=fila_existente, column=4, value=fecha_hora)

            ingreso_val = ws_reg.cell(row=fila_existente, column=3).value

            if ingreso_val:
                horas = (fecha_hora - ingreso_val).total_seconds() / 3600
                ws_reg.cell(row=fila_existente, column=5, value=round(horas, 2))

        else:
            # Caso de seguridad: no debería pasar porque la UI exige entrada primero
            ws_reg.append([estudiante, cultivo, None, fecha_hora, None])

    ws_dev.append([
        fecha_hora.strftime("%d/%m/%Y"),
        accion,
        device_id,
        estudiante
    ])

    recalcular_calendario(wb)

    wb.save(ARCHIVO_EXCEL)


def agregar_registro_manual(estudiante, cultivo, fecha, hora_ingreso, hora_salida):
    """Agrega directamente una fila a Registros. Se usa para cargar
    asistencias de días anteriores que no pasaron por la app (ej.
    registros en papel), sin pasar por la validación de dispositivo/GPS."""

    inicializar_excel()

    wb = load_workbook(ARCHIVO_EXCEL)
    ws_reg = wb["Registros"]

    ingreso_dt = datetime.combine(fecha, hora_ingreso)

    salida_dt = None
    horas = None

    if hora_salida:
        salida_dt = datetime.combine(fecha, hora_salida)
        horas = round((salida_dt - ingreso_dt).total_seconds() / 3600, 2)

    ws_reg.append([estudiante, cultivo, ingreso_dt, salida_dt, horas])

    recalcular_calendario(wb)

    wb.save(ARCHIVO_EXCEL)


# ============================================================
# PANEL DE ADMINISTRADOR (descargar Excel)
# ============================================================

HOJAS_VALIDAS = {"Registros", "_Dispositivos", "Calendario"}


def limpiar_hojas_obsoletas(wb):
    """Elimina cualquier hoja que no sea una de las que usa la app
    (por ejemplo, hojas creadas por versiones anteriores)."""

    for nombre_hoja in list(wb.sheetnames):

        if nombre_hoja not in HOJAS_VALIDAS:

            del wb[nombre_hoja]


CLAVE_ADMIN = "123"  # <-- clave administrador

with st.sidebar:

    st.subheader("🔒 Panel administrador")

    clave_ingresada = st.text_input(
        "Clave de acceso",
        type="password"
    )

    if clave_ingresada == CLAVE_ADMIN:

        inicializar_excel()

        wb_admin = load_workbook(ARCHIVO_EXCEL)

        # Quitar hojas de versiones anteriores que ya no se usan
        limpiar_hojas_obsoletas(wb_admin)

        # Se recalcula el Calendario con lo último registrado
        # justo en el momento de la descarga.
        recalcular_calendario(wb_admin)

        wb_admin.save(ARCHIVO_EXCEL)

        with open(ARCHIVO_EXCEL, "rb") as f:

            st.download_button(
                label="⬇️ Descargar asistencia.xlsx",
                data=f,
                file_name="asistencia.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        total_registros = wb_admin["Registros"].max_row - 1

        st.write(f"Filas de asistencia guardadas: **{total_registros}**")


        # ----------------------------------------------------
        # IMPORTAR REGISTROS DE DÍAS ANTERIORES
        # ----------------------------------------------------

        st.markdown("---")

        st.subheader("📥 Agregar registros anteriores")

        st.caption(
            "Para cargar asistencias que no pasaron por la app "
            "(por ejemplo, registros en papel de días pasados). "
            "Formato: Fecha = DD/MM/AAAA, Horas = HH:MM (24 horas). "
            "Deja 'Salida' vacía si esa persona solo tiene entrada."
        )

        tabla_importar = st.data_editor(
            pd.DataFrame({
                "Nombre": pd.Series(dtype="str"),
                "Cultivo": pd.Series(dtype="str"),
                "Fecha (DD/MM/AAAA)": pd.Series(dtype="str"),
                "Ingreso (HH:MM)": pd.Series(dtype="str"),
                "Salida (HH:MM)": pd.Series(dtype="str"),
            }),
            num_rows="dynamic",
            key="tabla_importar_manual",
            use_container_width=True
        )

        if st.button("➕ Agregar estos registros", use_container_width=True):

            filas_agregadas = 0
            filas_con_error = 0

            for _, fila in tabla_importar.iterrows():

                nombre_fila = str(fila.get("Nombre") or "").strip()
                cultivo_fila = str(fila.get("Cultivo") or "").strip()
                fecha_texto = str(fila.get("Fecha (DD/MM/AAAA)") or "").strip()
                ingreso_texto = str(fila.get("Ingreso (HH:MM)") or "").strip()
                salida_texto = str(fila.get("Salida (HH:MM)") or "").strip()

                fila_vacia = not nombre_fila and not fecha_texto

                if fila_vacia:
                    continue

                try:

                    if not nombre_fila or not fecha_texto or not ingreso_texto:
                        filas_con_error += 1
                        continue

                    fecha_normalizada = datetime.strptime(
                        fecha_texto, "%d/%m/%Y"
                    ).date()

                    ingreso_normalizado = datetime.strptime(
                        ingreso_texto, "%H:%M"
                    ).time()

                    salida_normalizada = None

                    if salida_texto:
                        salida_normalizada = datetime.strptime(
                            salida_texto, "%H:%M"
                        ).time()

                    agregar_registro_manual(
                        nombre_fila,
                        cultivo_fila,
                        fecha_normalizada,
                        ingreso_normalizado,
                        salida_normalizada
                    )

                    filas_agregadas += 1

                except ValueError:
                    filas_con_error += 1

            if filas_agregadas:
                st.success(f"✅ Se agregaron {filas_agregadas} registro(s).")
                st.info("Vuelve a entrar a este panel para descargar el Excel actualizado.")

            if filas_con_error:
                st.warning(
                    f"⚠️ {filas_con_error} fila(s) se omitieron por "
                    f"formato inválido. Revisa que Fecha sea "
                    f"DD/MM/AAAA y las horas HH:MM (ej. 07:00)."
                )

    elif clave_ingresada != "":

        st.error("Clave incorrecta")


# ============================================================
# TÍTULO
# ============================================================

st.title("📋 Asistencia Ciclo II 2026B")


st.write(
    "Seleccione su cultivo y registre su asistencia."
)


# ============================================================
# CULTIVOS
# ============================================================

cultivos = [
    "🥬 Acelga",
    "🌿 Alcachofa",
    "🔵 Arándano",
    "🌾 Avena",
    "🌱 Bulbo",
    "🥒 Calabacín",
    "🌿 Espárrago",
    "🌳 Feijoa",
    "🧅 Puerro",
    "🟠 Uchuva"
]


# ============================================================
# ESTUDIANTES
# ============================================================

estudiantes_por_cultivo = {

    "🥬 Acelga": [
        "Nancy Ximena Carrero",
        "Juan Esteban Huertas",
        "Daniel Cortés",
        "Gabriel Andres Guitierrez",
        "Liliana Maria Zea",
        "Paula Gabriela Riveros"
    ],

    "🌿 Alcachofa": [
        "Daniela Castillo",
        "Sergio Castellanos",
        "Alejandro Rojas",
        "Daniel Olivar",
        "Jhon Gómez",
        "jon Rodríguez"
    ],

    "🔵 Arándano": [
        "Angie Vanesa Lazo Sierra",
        "Harvey Steeven Ramos Puentes",
        "Sebastian Aparicio Guagua",
        "Jhon Alez Esquivel Benavides",
        "Alejandro Gil Cepeda",
        "Gustavo Andrés Garzón Pasachoa"
    ],

    "🌾 Avena": [
        "Ana Gabriela Rojas Gonzáles",
        "Emma Natalia Muñoz Jiménez",
        "Yuliana Henao Campuzano",
        "Jesus David Manzano",
        "Daniel Alberto Robles Sanchez",
        "Brayan Hair Recaman Montaño"
    ],

    "🌱 Bulbo": [
        "Marcos Fuerte Martinez",
        "Maria Paula Gonzales Urrego",
        "Becky Ortiz Rivas",
        "Cesar Augusto Ospino Nieto",
        "Luisa Rojas Rincón",
        "Stefany Julieth Torres Pinzón",
        "Ana Maria Valero Parra"
    ],

    "🥒 Calabacín": [
        "Edna Valentina Acosta Romero",
        "Juan Sebastian Camelo Garcia",
        "Luis Fernando Gaitan Pinto",
        "Sofia Molano Jara",
        "Luis Mario Pardo Fontecha",
        "David Felipe Robles Saavedra"
    ],

    "🌿 Espárrago": [
        "Sergio Castellanos Bello",
        "Karen Daniela Pineda Aldana",
        "Damián Nicolás Piracún Farfán",
        "Karen Dayana Gonzalés Prieto",
        "Brayan Stive Vanegas León"
    ],

    "🌳 Feijoa": [
        "Paula Andrea Rodríguez Mosquera",
        "Danna Kendris Castañeda Montealegre",
        "Juan Nicolás Aguilera Forero",
        "Javier Stiven Chávez Muñoz",
        "Jhon Ferney Derazo Fuelpaz",
        "Manuel Ricardo Vargas Alejo",
        "Darwin Alejandro Linares Cardenas"
    ],

    "🧅 Puerro": [
        "Ardila Penagos Valentina Vanessa Alexandra",
        "Cetina Díaz Juan Carlos",
        "Daza Juan Sebastian",
        "Reyes Junco Nicolás David",
        "Sánchez Villegas Luisa Fernanda",
        "Quevedo Viviana",
        "Quintero Camacho Daniel Andrés"
    ],

    "🟠 Uchuva": [
        "Angie Valentina Castillo Martín",
        "Tomás Leonardo Casas",
        "Andrés Felipe Aranguren",
        "Verónica Andrea Junco Trejos",
        "Raúl Mateo Vanegas",
        "Ingrid Tatiana Linares Anzola",
        "Daniel Alejandro Sánchez González"
    ]
}


# ============================================================
# VARIABLES DE SESIÓN
# ============================================================

if "cultivo_seleccionado" not in st.session_state:
    st.session_state.cultivo_seleccionado = None

if "estudiante_seleccionado" not in st.session_state:
    st.session_state.estudiante_seleccionado = None

if "accion" not in st.session_state:
    st.session_state.accion = None

if "solicitar_gps" not in st.session_state:
    st.session_state.solicitar_gps = False

if "device_id" not in st.session_state:
    st.session_state.device_id = None


# ============================================================
# FUNCIÓN PARA CALCULAR DISTANCIA
# ============================================================

def calcular_distancia(latitud, longitud):

    R = 6371000

    lat1 = radians(LATITUD_CAM)
    lat2 = radians(latitud)

    diferencia_latitud = radians(
        latitud - LATITUD_CAM
    )

    diferencia_longitud = radians(
        longitud - LONGITUD_CAM
    )

    a = (
        sin(diferencia_latitud / 2) ** 2
        +
        cos(lat1)
        * cos(lat2)
        * sin(diferencia_longitud / 2) ** 2
    )

    c = 2 * atan2(
        sqrt(a),
        sqrt(1 - a)
    )

    return R * c


# ============================================================
# OBTENER ID DE DISPOSITIVO (persistente en localStorage)
# ============================================================

device_id_actual = streamlit_js_eval(
    js_expressions="""
    (function() {
        let id = localStorage.getItem('asistencia_device_id');
        if (!id) {
            id = 'dev-' + Date.now() + '-' +
                 Math.random().toString(36).substring(2, 12);
            localStorage.setItem('asistencia_device_id', id);
        }
        return id;
    })()
    """,
    key="device_id_eval"
)

if device_id_actual:
    st.session_state.device_id = device_id_actual


# ============================================================
# PANTALLA 1
# SELECCIONAR CULTIVO
# ============================================================

if st.session_state.cultivo_seleccionado is None:

    st.subheader("🌱 Seleccione su cultivo")

    st.markdown("---")

    col1, col2 = st.columns(2)

    for i, cultivo in enumerate(cultivos):

        if i % 2 == 0:

            with col1:

                if st.button(
                    cultivo,
                    use_container_width=True,
                    key=f"cultivo_{i}"
                ):

                    st.session_state.cultivo_seleccionado = cultivo

                    st.rerun()

        else:

            with col2:

                if st.button(
                    cultivo,
                    use_container_width=True,
                    key=f"cultivo_{i}"
                ):

                    st.session_state.cultivo_seleccionado = cultivo

                    st.rerun()


# ============================================================
# PANTALLA 2
# SELECCIONAR ESTUDIANTE
# ============================================================

elif st.session_state.estudiante_seleccionado is None:

    cultivo = st.session_state.cultivo_seleccionado

    st.subheader(
        f"{cultivo}"
    )

    st.write(
        "Seleccione su nombre."
    )

    st.markdown("---")

    estudiantes = estudiantes_por_cultivo[cultivo]

    for i, estudiante in enumerate(estudiantes):

        if st.button(
            estudiante,
            use_container_width=True,
            key=f"estudiante_{i}"
        ):

            st.session_state.estudiante_seleccionado = estudiante

            st.rerun()


    st.markdown("---")

    if st.button(
        "⬅️ Cambiar de cultivo",
        use_container_width=True
    ):

        st.session_state.cultivo_seleccionado = None

        st.rerun()


# ============================================================
# PANTALLA 3
# ESTUDIANTE
# ============================================================

else:

    cultivo = st.session_state.cultivo_seleccionado

    estudiante = st.session_state.estudiante_seleccionado


    st.subheader(
        f"👤 {estudiante}"
    )

    st.write(
        f"🌱 Cultivo: {cultivo}"
    )

    st.markdown("---")


    # ========================================================
    # BOTONES ENTRADA / SALIDA
    # ========================================================

    if not st.session_state.solicitar_gps:

        st.write(
            "Seleccione la acción que desea registrar:"
        )

        fecha_hoy = ahora_bogota().strftime("%d/%m/%Y")

        entrada_hecha = dispositivo_ya_registro(
            st.session_state.device_id, fecha_hoy, "Entrada"
        )

        salida_hecha = dispositivo_ya_registro(
            st.session_state.device_id, fecha_hoy, "Salida"
        )

        entrada_es_de_este_estudiante = dispositivo_ya_registro(
            st.session_state.device_id, fecha_hoy, "Entrada", estudiante
        )

        col1, col2 = st.columns(2)


        # ----------------------------------------------------
        # ENTRADA
        # ----------------------------------------------------

        with col1:

            if entrada_hecha:

                st.button(
                    "🟢 ENTRAR",
                    use_container_width=True,
                    disabled=True
                )

                st.caption(
                    "✅ Este dispositivo ya registró entrada hoy"
                )

            else:

                if st.button(
                    "🟢 ENTRAR",
                    use_container_width=True
                ):

                    st.session_state.accion = "Entrada"

                    st.session_state.solicitar_gps = True

                    st.rerun()


        # ----------------------------------------------------
        # SALIDA
        # ----------------------------------------------------

        with col2:

            if salida_hecha:

                st.button(
                    "🔴 SALIR",
                    use_container_width=True,
                    disabled=True
                )

                st.caption(
                    "✅ Este dispositivo ya registró salida hoy"
                )

            elif not entrada_hecha:

                st.button(
                    "🔴 SALIR",
                    use_container_width=True,
                    disabled=True
                )

                st.caption(
                    "⚠️ Primero debes registrar la entrada"
                )

            elif not entrada_es_de_este_estudiante:

                st.button(
                    "🔴 SALIR",
                    use_container_width=True,
                    disabled=True
                )

                st.caption(
                    "⚠️ La entrada de este dispositivo fue para "
                    "otro estudiante. Seleccione a esa persona "
                    "para poder registrar la salida."
                )

            else:

                if st.button(
                    "🔴 SALIR",
                    use_container_width=True
                ):

                    st.session_state.accion = "Salida"

                    st.session_state.solicitar_gps = True

                    st.rerun()


    # ========================================================
    # SOLICITAR GPS
    # ========================================================

    else:

        accion = st.session_state.accion


        if accion == "Entrada":

            st.subheader(
                "🟢 Registrando entrada"
            )

        else:

            st.subheader(
                "🔴 Registrando salida"
            )


        st.info(
            "📍 Obteniendo ubicación..."
        )


        # ====================================================
        # JAVASCRIPT PARA OBTENER GPS
        # ====================================================

        ubicacion = streamlit_js_eval(
            js_expressions="""
            new Promise((resolve, reject) => {

                navigator.geolocation.getCurrentPosition(

                    position => {

                        resolve({

                            latitude:
                            position.coords.latitude,

                            longitude:
                            position.coords.longitude,

                            accuracy:
                            position.coords.accuracy

                        });

                    },

                    error => {

                        resolve({

                            error:
                            error.message

                        });

                    },

                    {

                        enableHighAccuracy: true,

                        timeout: 10000,

                        maximumAge: 0

                    }

                );

            })
            """,

            key="obtener_gps"
        )


        # ====================================================
        # PROCESAR GPS
        # ====================================================

        if ubicacion is not None:


            # ------------------------------------------------
            # ERROR GPS
            # ------------------------------------------------

            if "error" in ubicacion:

                st.error(
                    "❌ No fue posible obtener su ubicación."
                )

                st.write(
                    ubicacion["error"]
                )


            # ------------------------------------------------
            # GPS CORRECTO
            # ------------------------------------------------

            else:

                latitud = ubicacion["latitude"]

                longitud = ubicacion["longitude"]

                precision = ubicacion["accuracy"]


                # --------------------------------------------
                # CALCULAR DISTANCIA
                # --------------------------------------------

                distancia = calcular_distancia(
                    latitud,
                    longitud
                )


                st.metric(
                    "Distancia al CAM",
                    f"{distancia:.1f} metros"
                )


                st.write(
                    f"Precisión GPS: "
                    f"{precision:.1f} metros"
                )


                # ============================================
                # UBICACIÓN AUTORIZADA
                # ============================================

                if distancia <= RADIO_PERMITIDO:


                    fecha_hora = ahora_bogota()

                    fecha_hoy = fecha_hora.strftime("%d/%m/%Y")

                    device_id = st.session_state.device_id


                    # Revalidación final por si dos pestañas
                    # del mismo dispositivo intentaron registrar
                    # al mismo tiempo, o si intentan dar salida
                    # a un estudiante distinto al que dio entrada.

                    ya_registrado = dispositivo_ya_registro(
                        device_id, fecha_hoy, accion
                    )

                    entrada_de_otro_estudiante = (
                        accion == "Salida"
                        and not dispositivo_ya_registro(
                            device_id, fecha_hoy, "Entrada", estudiante
                        )
                    )

                    if ya_registrado:

                        st.error(
                            f"❌ Este dispositivo ya registró "
                            f"la {accion.lower()} de hoy."
                        )

                    elif entrada_de_otro_estudiante:

                        st.error(
                            "❌ La entrada de este dispositivo fue "
                            "para otro estudiante. No se puede "
                            "registrar la salida de "
                            f"{estudiante}."
                        )

                    else:

                        guardar_registro(
                            estudiante,
                            cultivo,
                            accion,
                            distancia,
                            device_id,
                            fecha_hora
                        )


                        st.success(
                            "🟢 Ubicación autorizada"
                        )


                        st.success(
                            f"✅ {accion} registrada correctamente"
                        )


                        st.write(
                            f"👤 **Estudiante:** "
                            f"{estudiante}"
                        )


                        st.write(
                            f"🌱 **Cultivo:** "
                            f"{cultivo}"
                        )


                        st.write(
                            f"📅 **Fecha:** "
                            f"{fecha_hora.strftime('%d/%m/%Y')}"
                        )


                        st.write(
                            f"⏰ **Hora:** "
                            f"{fecha_hora.strftime('%H:%M:%S')}"
                        )


                # ============================================
                # UBICACIÓN NO AUTORIZADA
                # ============================================

                else:


                    st.error(
                        "🔴 Ubicación no autorizada"
                    )


                    st.write(
                        f"Se encuentra a "
                        f"**{distancia:.1f} metros** "
                        f"del CAM."
                    )


                    st.write(
                        f"El máximo permitido es "
                        f"**{RADIO_PERMITIDO} metros**."
                    )


    # ========================================================
    # VOLVER
    # ========================================================

    st.markdown("---")


    if st.button(
        "⬅️ Volver a lista de estudiantes",
        use_container_width=True
    ):

        st.session_state.estudiante_seleccionado = None

        st.session_state.accion = None

        st.session_state.solicitar_gps = False

        st.rerun()